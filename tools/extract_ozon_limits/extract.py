#!/usr/bin/env python3
"""Extract Ozon category field limits from seller xlsx templates (configs sheet base64 JSON)."""

from __future__ import annotations

import argparse
import base64
import csv
import glob
import json
import os
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_SRC = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - Allure City, Inc",
    "ОП_E-COMMERCE",
    "Автоматизация",
    "Шаблоны для SEO-Ozon",
)

EXPECTED_CATEGORIES = [
    "Парфюмерия",
    "Свеча",
    "Соль для ванны",
    "Косметика для ухода",
    "Косметика для волос",
    "Ароматы для дома",
    "Средство после бритья",
    "Средства для гигиены тела",
]

# Human rules from Техздание SEO OZON.docx (not always in base64 MaxValue)
MANUAL_RULES = {
    "title": {
        "field_name_ru": "Название товара",
        "max_length": 200,
        "notes": "≤27 симв/слово; без CAPS; без HTML; см. Техздание SEO OZON",
    },
    "annotation": {
        "field_id": "4191",
        "field_name_ru": "Аннотация",
        "allowed_html": ["br", "ul", "li"],
        "notes": "HTML только br/ul/li; см. Техздание SEO OZON",
    },
}


def find_xlsx_files(src: str) -> list[str]:
    if os.path.isdir(src):
        patterns = [os.path.join(src, "*.xlsx"), os.path.join(src, "**", "*.xlsx")]
        files: list[str] = []
        for pat in patterns:
            files.extend(glob.glob(pat, recursive=True))
        return sorted(set(files))
    return sorted(glob.glob(src, recursive=True))


def _collect_configs_blob(ws) -> str:
    parts: list[str] = []
    for row in ws.iter_rows(max_row=40, max_col=30):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()
            if len(s) >= 50 and re.fullmatch(r"[A-Za-z0-9+/=\s]+", s):
                parts.append(s.replace("\n", "").replace("\r", ""))
    return "".join(parts)


def decode_configs_sheet(wb_path: str) -> dict[str, Any]:
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    sheet_names = {s.lower(): s for s in wb.sheetnames}
    if "configs" not in sheet_names:
        wb.close()
        raise ValueError("sheet 'configs' not found")
    ws = wb[sheet_names["configs"]]
    blob = _collect_configs_blob(ws)
    wb.close()
    if not blob:
        raise ValueError("empty configs blob")
    raw = base64.b64decode(blob)
    return json.loads(raw.decode("utf-8"))


def _allowed_values_count(attr: dict[str, Any]) -> int:
    for key in ("Dictionary", "Options", "Values", "AllowedValues"):
        val = attr.get(key)
        if isinstance(val, list):
            return len(val)
        if isinstance(val, dict):
            opts = val.get("Options") or val.get("values") or val.get("Values")
            if isinstance(opts, list):
                return len(opts)
            if isinstance(opts, dict):
                return len(opts)
    return 0


def extract_fields(data: dict[str, Any]) -> list[dict[str, Any]]:
    category = data.get("name") or data.get("Name") or ""
    attrs = data.get("attributes") or data.get("Attributes") or {}
    rows: list[dict[str, Any]] = []

    for field_id, attr in attrs.items():
        if not isinstance(attr, dict):
            continue
        name = attr.get("Name") or attr.get("LongName") or attr.get("Label") or ""
        rows.append(
            {
                "ozon_category": category,
                "field_id": str(field_id),
                "field_name_ru": name,
                "data_type": attr.get("Type") or attr.get("type") or "",
                "max_length": attr.get("MaxValue") or attr.get("max_length") or "",
                "is_required": attr.get("IsRequired", attr.get("is_required", "")),
                "is_collection": attr.get("IsCollection", attr.get("is_collection", "")),
                "allowed_values_count": _allowed_values_count(attr),
                "notes": "",
            }
        )

    # Inject manual SEO rules when category parsed
    if category:
        rows.append(
            {
                "ozon_category": category,
                "field_id": "TITLE_MANUAL",
                "field_name_ru": MANUAL_RULES["title"]["field_name_ru"],
                "data_type": "String",
                "max_length": MANUAL_RULES["title"]["max_length"],
                "is_required": True,
                "is_collection": False,
                "allowed_values_count": 0,
                "notes": MANUAL_RULES["title"]["notes"],
            }
        )
        rows.append(
            {
                "ozon_category": category,
                "field_id": MANUAL_RULES["annotation"]["field_id"],
                "field_name_ru": MANUAL_RULES["annotation"]["field_name_ru"],
                "data_type": "multiline",
                "max_length": "",
                "is_required": False,
                "is_collection": False,
                "allowed_values_count": 0,
                "notes": "allowed_html=" + ",".join(MANUAL_RULES["annotation"]["allowed_html"]),
            }
        )
    return rows


def build_limits_dict(all_rows: list[dict[str, Any]]) -> dict[str, dict[str, dict[str, Any]]]:
    out: dict[str, dict[str, dict[str, Any]]] = {}
    for row in all_rows:
        cat = row["ozon_category"]
        fid = row["field_id"]
        out.setdefault(cat, {})[fid] = {
            k: row[k]
            for k in row
            if k not in ("ozon_category",)
        }
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract Ozon limits from xlsx templates")
    parser.add_argument("--src", default=DEFAULT_SRC, help="Folder with Ozon xlsx templates")
    parser.add_argument(
        "--out-dir",
        default=str(Path(__file__).resolve().parent),
        help="Output directory for json/csv",
    )
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Fallback: glob OneDrive if default path missing
    src = args.src
    if not os.path.isdir(src):
        alt = glob.glob(
            os.path.join(
                os.path.expanduser("~"),
                "OneDrive - Allure City, Inc",
                "**",
                "*SEO-Ozon*",
            ),
            recursive=True,
        )
        alt_dirs = [p for p in alt if os.path.isdir(p)]
        if alt_dirs:
            src = alt_dirs[0]
            print(f"WARN: default src missing, using discovered: {src}", file=sys.stderr)

    files = find_xlsx_files(src)
    log_lines: list[str] = []
    all_rows: list[dict[str, Any]] = []
    found_categories: set[str] = set()

    for path in files:
        base = os.path.basename(path)
        try:
            data = decode_configs_sheet(path)
            cat = data.get("name") or base
            found_categories.add(cat)
            rows = extract_fields(data)
            all_rows.extend(rows)
            log_lines.append(f"OK  {base} -> category '{cat}', attributes={len(rows)-2}")
        except Exception as exc:  # noqa: BLE001
            log_lines.append(f"ERR {base}: {exc}")

    missing = [c for c in EXPECTED_CATEGORIES if c not in found_categories]
    for cat in missing:
        log_lines.append(f"MISSING category (no xlsx): {cat}")

    limits = build_limits_dict(all_rows)
    json_path = out_dir / "ozon_limits.json"
    csv_path = out_dir / "ozon_limits.csv"

    with json_path.open("w", encoding="utf-8") as f:
        json.dump(limits, f, ensure_ascii=False, indent=2)

    fieldnames = [
        "ozon_category",
        "field_id",
        "field_name_ru",
        "data_type",
        "max_length",
        "is_required",
        "is_collection",
        "allowed_values_count",
        "notes",
    ]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"Processed xlsx files: {len(files)}")
    print(f"Categories recognized: {len(found_categories)}")
    print(f"Missing expected count: {len(missing)}")
    print(f"Total CSV rows: {len(all_rows)}")
    print(f"JSON size: {json_path.stat().st_size} bytes")
    print("Log:")
    for line in log_lines:
        print(" ", line)
    print(f"Wrote {json_path}")
    print(f"Wrote {csv_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
